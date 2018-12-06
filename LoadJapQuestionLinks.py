import urllib.error, urllib.request
import urllib.parse
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook


# read reading exercises
nrOfExercisesReading = [9,12,18,10,11]
nrOfExercisesGrammar = [25,25,29,30,26]
nrOfExercisesKanji = [21,21,31,20,20]
nrOfExercisesVocabulary = [22,26,21,31,24]

hdr = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
       'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.0.7) Gecko/2009021910 Firefox/3.0.7'}


def formExerciseURL(level, nr, type):
    return "https://japanesetest4you.com/japanese-language-proficiency-test-jlpt-n" + str(level) + "-" + type + "-" + str(nr) + "/"

grammarPages = []
readingPages = []
kanjiPages = []
vocabularyPages = []

for level in range(1,5+1):
    for nr in range(1,nrOfExercisesGrammar[level-1]+1):
        grammarPages.append([level, nr, formExerciseURL(level, nr, "grammar-exercise")])
    for nr in range(1,nrOfExercisesKanji[level-1]+1):
        readingPages.append([level, nr, formExerciseURL(level, nr, "kanji-exercise")])
    for nr in range(1,nrOfExercisesReading[level-1]+1):
        kanjiPages.append([level, nr, formExerciseURL(level, nr, "reading-exercise")])
    for nr in range(1,nrOfExercisesVocabulary[level-1]+1):
        vocabularyPages.append([level, nr, formExerciseURL(level, nr, "vocabulary-exercise")])

def readQuestionsFromURL(url):
    request = urllib.request.Request(url, None, hdr)
    response = urllib.request.urlopen(request)
    page = response.read()
    beu = BeautifulSoup(page)
    form = beu.findAll("form")[0]
    questionsRaw = form.findAll("p")

    questions = []


    for q in questionsRaw:
        q2 = str(q).replace("input", "%") + "%"
        if (str(q).find("Advertisement")>=0):
            continue
        if (str(q).find("Submit Quiz")>=0):
            continue
        if (str(q).find("input")<0):
            continue
        questionText = q.contents[0]

        answers = re.findall('value=\".\"/>([^%]*)%', str(q2))
        questions.append([questionText,answers])
    return questions

def readQuestionsFromURLComplete(url):
    request = urllib.request.Request(url, None, hdr)
    response = urllib.request.urlopen(request)
    page = response.read()
    beu = BeautifulSoup(page)
    form = beu.findAll("form")[0]
    x = str(form)

    x = re.sub("<p>", "", x)
    x = re.sub("</p>", "", x)
    x = re.sub("<br/>", "", x)
    x = re.sub("<span[^>]*>", "{{", x)
    x = re.sub("</span[^>]*>", "}}", x)
    x = re.sub("<input[^>]*>", "( )", x)
    x = re.sub("<form[^>]*>", "", x)
    x = re.sub("<div[^>]*>", "", x)
    x = re.sub("<script[^>]*>", "", x)
    x = re.sub("</script[^>]*>", "", x)
    s = re.compile("<!--.*-->", re.DOTALL)
    x = re.sub(s, "",x)

    x = re.sub("\n\n", "\n", x)
    x = re.sub("\n\n", "\n", x)

    s = re.compile("Advertisement[^\n]*\n", re.DOTALL)
    x = re.sub(s, "", x)

    re.findall("<!--.*-->", x, re.DOTALL)

    print(x)


    form.contents[3]

    questionsRaw = form.findAll("p")

    questions = []


    for q in questionsRaw:
        q2 = str(q).replace("input", "%") + "%"
        if (str(q).find("Advertisement")>=0):
            continue
        if (str(q).find("Submit Quiz")>=0):
            continue
        if (str(q).find("input")<0):
            continue
        questionText = q.contents[0]

        answers = re.findall('value=\".\"/>([^%]*)%', str(q2))
        questions.append([questionText,answers])
    return questions

def getExerciseSolution2(url,nrOfQuestions):
    requesturl = url
    postData  = {}

    trueValues = [0] * nrOfQuestions
    for sol in range(1,5):
        for i in range(nrOfQuestions):
            postData["quest" + str(i+1)] = str(sol)
        postData["submit"] = "Submit+Quiz"
        request = urllib.request.Request(url,urllib.parse.urlencode(postData).encode("utf-8"), hdr)
        response = urllib.request.urlopen(request)
        page = response.read()
        beu = BeautifulSoup(page)
        cont = beu.findAll("div", {"class":"entry clearfix"})[0]
        solutionString = str(cont)
        solutions = re.findall("(green|red)",solutionString)

        for x in range(solutions.__len__()):
            if solutions[x]=="green":
                trueValues[x] = sol

    return trueValues


def cleanExerciseText(exercises):
    for q in range(exercises.__len__()):
        for i in range(4,9):
            exercises[q][i] = exercises[q][i].replace("<br/>\\n<", "")
            exercises[q][i] = exercises[q][i].replace("<br/>", "")
            exercises[q][i] = exercises[q][i].replace("</p>", "")
            exercises[q][i] = exercises[q][i].replace("</br>", "")
            exercises[q][i] = exercises[q][i].replace("</br>", "")
            exercises[q][i] = exercises[q][i].replace("<", "")
            exercises[q][i] = exercises[q][i].strip()
    return exercises

def writeExercises(exercises,name):
    wb = Workbook()
    ws = wb.active
    for e in range(exercises.__len__()):
        ws['A'+str(e+1)] = exercises[e][0]
        ws['B'+str(e+1)] = exercises[e][1]
        ws['C'+str(e+1)] = exercises[e][2]
        ws['D'+str(e+1)] = exercises[e][3]
        ws['E'+str(e+1)] = exercises[e][4]
        ws['F'+str(e+1)] = exercises[e][5]
        ws['G'+str(e+1)] = exercises[e][6]
        ws['H'+str(e+1)] = exercises[e][7]
        ws['I'+str(e+1)] = exercises[e][8]
        ws['J'+str(e+1)] = exercises[e][9]
    wb.save(name)




grammarExercises = []
for i in range(0,grammarPages.__len__()):
    try:
        u = grammarPages[i]
        print(i)
        questions = readQuestionsFromURL(u[2])
        solutions = getExerciseSolution2(u[2], questions.__len__())
        for q in range(questions.__len__()):
            grammarExercises.append([ u[0], u[1], u[2], q, questions[q][0], questions[q][1][0], questions[q][1][1], questions[q][1][2], questions[q][1][3], solutions[q]  ])
    except:
        print("Exception")
# clean text
grammarExercises = cleanExerciseText(grammarExercises)
writeExercises(grammarExercises, "grammarExercises.xlsx")



readingExercises = []
for i in range(0,readingPages.__len__()):
for i in range(0,10):
    try:
        u = readingPages[i]
        print(i)
        questions = readQuestionsFromURL(u[2])
        solutions = getExerciseSolution2(u[2], questions.__len__())
        for q in range(questions.__len__()):
            readingExercises.append([ u[0], u[1], u[2], q, questions[q][0], questions[q][1][0], questions[q][1][1], questions[q][1][2], questions[q][1][3], solutions[q]  ])
    except:
        print("Exception")







