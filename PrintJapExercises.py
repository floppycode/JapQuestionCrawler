import win32api
import win32print
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import uuid

STARTNUMBER = 5

if uuid.getnode() == 345052807176:
    # create new temporary file
    wb_out = load_workbook('D:/Dropbox/12Projects/JapEx/grammarExercisePrintForm.xlsx')
    ws_out = wb_out.active
    # load feed
    wb_feed = load_workbook('D:/Dropbox/12Projects/JapEx/grammarFeed.xlsx')
    ws_feed = wb_feed.active
elif uuid.getnode() == 23414543623:
    # create new temporary file
    wb_out = load_workbook('C:/Users/flopp/Dropbox/12Projects/JapEx/grammarExercisePrintForm.xlsx')
    ws_out = wb_out.active
    # load feed
    wb_feed = load_workbook('C:/Users/flopp/Dropbox/12Projects/JapEx/grammarFeed.xlsx')
    ws_feed = wb_feed.active
else:
    quit()

# kopiere in vorlagae
def setExerciseNr(targetSlot,nr):
    base = (targetSlot-1)*5
    ws_out['A'+str(base+1)].value = str(ws_feed['E'+str(nr)].value)
    ws_out['A'+str(base+2)].value = '( )' + str(ws_feed['F'+str(nr)].value)
    ws_out['A'+str(base+3)].value = '( )' + str(ws_feed['G'+str(nr)].value)
    ws_out['A'+str(base+4)].value = '( )' + str(ws_feed['H'+str(nr)].value)
    ws_out['A'+str(base+5)].value = '( )' + str(ws_feed['I'+str(nr)].value)

def setAlignment(nr):
    for i in range(1,nr+1):
        ws_out['A'+str(i)].alignment = Alignment(wrapText=True)

def createSolution(nr):
    sol = ""
    for i in range(STARTNUMBER,STARTNUMBER+nr):
        sol = sol + str(ws_feed["J"+str(i)].value)
    return sol

setExerciseNr(1, STARTNUMBER)
setExerciseNr(2, STARTNUMBER+1)
setExerciseNr(3, STARTNUMBER+2)
setExerciseNr(4, STARTNUMBER+3)
setAlignment(5*4)
sol = createSolution(4)
ws_out['A22'].value = sol + "     " +str(STARTNUMBER)
ws_out["A22"].font = ws_out["A22"].font.copy(strike=True, sz= 6 )


if uuid.getnode() == 345052807176:
    wb_out.save("D:/Dropbox/12Projects/JapEx/tempGrammarForm.xlsx")
    filename = "D:/Dropbox/12Projects/JapEx/tempGrammarForm.xlsx"
elif uuid.getnode() == 23452346123:
    wb_out.save("C:/Users/flopp/Dropbox/12Projects/JapEx/tempGrammarForm.xlsx")
    filename = "C:/Users/flopp/Dropbox/12Projects/JapEx/tempGrammarForm.xlsx"


win32api.ShellExecute (
  0,
  "printto",
  filename,
  '"%s"' % win32print.GetDefaultPrinter (),
  ".",
  0
)








